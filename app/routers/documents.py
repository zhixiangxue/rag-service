"""Document API endpoints."""
from fastapi import APIRouter, HTTPException, UploadFile, File, Form, Body, Query
from fastapi.responses import FileResponse, StreamingResponse
from typing import List, Optional, Dict, Any
import asyncio
import os
import json
import sqlite3
import tempfile
from pathlib import Path

from zag.utils.hash import calculate_file_hash
from ..utils.s3 import get_s3_object_info, download_file_from_s3_async
from ..database import get_connection, now, generate_id
from ..repositories import DatasetRepository, DocumentRepository, TaskRepository, DependencyRepository
from ..domain.deps import Rule, DependencySource
from ..schemas import (
    DocumentResponse,
    ApiResponse,
    MessageResponse,
    TaskResponse,
    ProcessingMode,
    ReaderType,
    _extract_tags,
    LocatePageRequest,
    LocatePageResult,
    LocatePageResponse,
)
from ..storage import get_storage
from ..constants import TaskStatus, DocumentStatus
from .. import config
from ..worker import later

router = APIRouter(prefix="/datasets/{dataset_id}/documents", tags=["documents"])

# ---------------------------------------------------------------------------
# Module-level diskcache singleton for locate_pages
# Stores (full_text: str, page_positions: list[(start, end, page_num)]) per doc_id.
# Pre-warm with: python playground/prewarm_locate_cache.py
# Initialized once at module load time — no lazy init, no race condition.
# ---------------------------------------------------------------------------

def _init_locate_cache():
    """Open the diskcache at module load time. Returns None on error."""
    try:
        import diskcache
        cache_dir = Path(config.LOCATE_CACHE_DIR)
        cache_dir.mkdir(parents=True, exist_ok=True)
        cache = diskcache.Cache(str(cache_dir), size_limit=int(20e9))
        print(f"[locate] diskcache opened: {cache_dir}  entries={len(cache)}")
        return cache
    except Exception as exc:
        print(f"[locate] diskcache unavailable, falling back to fitz: {exc}")
        return None


_locate_cache = _init_locate_cache()


def _get_locate_cache():
    """Return the module-level diskcache.Cache singleton."""
    return _locate_cache


def _extract_pdf_text_and_positions(file_path: str):
    """
    Extract normalized text and page character-positions from a PDF via fitz.

    Replicates the exact same normalization used in find_pages_in_document so that
    cached data is byte-for-byte equivalent to what the live path would produce.

    Returns:
        (full_text, page_positions) on success, or None on failure/scanned PDF.
        page_positions: list of (page_start_char, page_end_char, page_number_1based)
    """
    try:
        import fitz
        from zag.utils.page_inference import normalize_text

        doc = fitz.open(file_path)
        full_text = ""
        page_positions = []
        current = 0
        try:
            for page in doc:
                norm = normalize_text(page.get_text())
                page_start = current
                page_end = current + len(norm)
                page_positions.append((page_start, page_end, page.number + 1))
                full_text += ("\n" if full_text else "") + norm
                current = page_end + 1

                # Early bail-out for scanned PDFs (< 50 chars/page after 3 pages)
                if page.number == 2 and len(full_text) / 3 < 50:
                    return None
        finally:
            doc.close()

        return full_text, page_positions
    except Exception:
        return None


def _search_in_text(
    full_text: str,
    page_positions: list,
    text_start: str,
    text_end: str | None,
) -> tuple[list[int] | None, bool]:
    """
    Search for text_start / text_end inside pre-extracted full_text.

    Returns:
        (page_numbers, found) — same contract as find_pages_in_document.
    """
    from zag.utils.page_inference import fuzzy_find_start, normalize_text

    norm_start = normalize_text(text_start)
    norm_end = normalize_text(text_end) if text_end else None

    start_pos = fuzzy_find_start(norm_start, full_text, start_from=0, threshold=0.85)
    if start_pos is None:
        return None, False

    # Reject ambiguous matches: if text_start appears more than once, the location
    # is undetermined — returning the first hit would likely be wrong.
    duplicate = fuzzy_find_start(
        norm_start, full_text,
        start_from=start_pos + 1,
        threshold=0.85,
    )
    if duplicate is not None:
        return None, False

    if norm_end:
        found_end = fuzzy_find_start(
            norm_end, full_text,
            start_from=start_pos + len(norm_start),
            threshold=0.85,
            max_search_range=100_000,
        )
        end_pos = (found_end + len(norm_end)) if found_end is not None else (start_pos + len(norm_start))
    else:
        end_pos = start_pos + len(norm_start)

    pages = sorted(
        pn for ps, pe, pn in page_positions
        if not (end_pos <= ps or start_pos >= pe)
    )
    if not pages:
        return [], False

    # Reject implausibly wide spans — likely a false-positive start match.
    if len(pages) > 10:
        return None, False

    return pages, True


def _build_file_url(stored_path: str, base_dir_str: Optional[str] = None) -> Optional[str]:
    """Build file_url from stored file path (s3:// or local)."""
    if not stored_path:
        return None
    if stored_path.startswith("s3://"):
        return stored_path
    # Local file: pure string processing, no filesystem calls
    norm_path = stored_path.replace("\\", "/")
    if base_dir_str is None:
        base_dir_str = Path(get_storage().base_dir).resolve().as_posix()
    # Make absolute if relative
    if not norm_path.startswith("/"):
        norm_path = Path.cwd().as_posix() + "/" + norm_path
    # Strip base_dir prefix to get relative path
    base = base_dir_str.rstrip("/") + "/"
    if norm_path.startswith(base):
        rel_path = norm_path[len(base):]
        return f"http://{config.API_PUBLIC_HOST}:{config.API_PORT}/files/{rel_path}"
    return None


def _validate_metadata(metadata: dict) -> dict:
    """
    Validate metadata dict structure.

    Args:
        metadata: Metadata dict

    Returns:
        The validated metadata dict

    Raises:
        ValueError: If metadata is invalid or guideline/overlays values are not allowed
    """
    if not isinstance(metadata, dict):
        raise ValueError("Metadata must be a JSON object")

    # Validate guideline and overlays fields
    VALID_GUIDELINES = {"FannieMae", "FreddieMac", "VA", "USDA", "FHA"}

    if "guideline" in metadata:
        guideline_val = metadata["guideline"]
        if guideline_val not in VALID_GUIDELINES:
            raise ValueError(
                f'Invalid guideline value: "{guideline_val}". '
                f"Must be one of: {sorted(VALID_GUIDELINES)}"
            )

    if "overlays" in metadata:
        overlays = metadata["overlays"]
        if not isinstance(overlays, list):
            raise ValueError('overlays must be an array')
        for item in overlays:
            if item not in VALID_GUIDELINES:
                raise ValueError(
                    f'Invalid overlays value: "{item}". '
                    f"Must be one of: {sorted(VALID_GUIDELINES)}"
                )

    return metadata


def _create_document_record(
    dataset_id: str,
    file_path: str,
    filename: str,
    metadata: Optional[dict] = None,
    s3_url: Optional[str] = None,
) -> dict:
    """
    Create document record after a local file is saved.
    Handles dataset validation, hash calculation, duplicate check, and database insert.

    Args:
        dataset_id: Dataset ID
        file_path: Local path to the saved file (used for hash/size calculation)
        filename: Original filename
        metadata: Validated metadata dict
        s3_url: If provided, stored as file_path in DB instead of local path

    Returns:
        dict with doc_id, file_name, file_path, file_hash, and status info

    Raises:
        ValueError: If metadata is invalid
        HTTPException: If dataset not found (404)
    """
    # The path actually stored in DB
    db_path = s3_url if s3_url else file_path
    conn = get_connection()
    doc_repo = DocumentRepository(conn)

    try:
        # Check if dataset exists
        if not DatasetRepository(conn).exists(dataset_id):
            raise HTTPException(status_code=404, detail="Dataset not found")

        # Calculate file hash using zag's utility on saved file
        file_hash = calculate_file_hash(file_path)

        # Check for duplicate: same dataset + same file_hash
        existing_doc = doc_repo.get_by_hash(dataset_id, file_hash)

        if existing_doc:
            # Duplicate: update file_path to latest and return
            doc_repo.update_file_path(existing_doc["id"], db_path, now())
            return {
                "dataset_id": str(dataset_id),
                "doc_id": str(existing_doc["id"]),
                "file_name": existing_doc["file_name"],
                "file_hash": file_hash,
                "is_duplicate": True
            }

        # Extract workspace directory (parent of file)
        workspace_dir = os.path.dirname(db_path)

        file_size = os.path.getsize(file_path)
        file_type = filename.split(".")[-1] if "." in filename else "unknown"

        # Serialize metadata to JSON for storage
        if metadata is None:
            raise ValueError(
                "metadata is required. Please provide a JSON object. "
                'Recommended fields: {"lender": "xxx", "guideline": "FannieMae|...", '
                '"overlays": ["xxx"], "tags": ["xxx"]}'
            )
        metadata_json = json.dumps(metadata)

        timestamp = now()
        # Use file_hash as doc_id to ensure same content gets same ID
        doc_id = file_hash

        # Create Document record with file_hash and metadata
        # Use INSERT OR IGNORE to handle race conditions
        try:
            doc_repo.create(
                doc_id, dataset_id, filename, db_path, workspace_dir,
                file_size, file_type, file_hash, metadata_json,
                DocumentStatus.PROCESSING, timestamp
            )
            return {
                "dataset_id": str(dataset_id),
                "doc_id": str(doc_id),
                "file_name": filename,
                "file_path": db_path,
                "file_hash": file_hash,
                "is_duplicate": False
            }
        except sqlite3.IntegrityError:
            # Race condition: another request inserted the same file first, just return duplicate
            conn.rollback()
            existing_doc = doc_repo.get_by_hash(dataset_id, file_hash)
            return {
                "dataset_id": str(dataset_id),
                "doc_id": str(existing_doc["id"]),
                "file_name": existing_doc["file_name"],
                "file_hash": file_hash,
                "is_duplicate": True
            }
    finally:
        conn.close()


def _cache_pdf_to_files_dir(src_path: str, doc_id: str) -> None:
    """Copy uploaded PDF to PDF_FILES_DIR/{doc_id}.pdf for locate_pages lookup.

    Only copies .pdf files. Skips silently if file already exists or on any error
    so the main upload flow is never affected.

    Args:
        src_path: Local path to the source PDF file
        doc_id: Document ID used as the target filename
    """
    import shutil
    if not src_path.lower().endswith(".pdf"):
        return
    try:
        dest_dir = Path(config.PDF_FILES_DIR)
        dest_dir.mkdir(parents=True, exist_ok=True)
        dest = dest_dir / f"{doc_id}.pdf"
        if not dest.exists():
            shutil.copy2(src_path, dest)
    except Exception as e:
        print(f"[WARN] Failed to cache PDF to PDF_FILES_DIR for doc_id {doc_id}: {e}")


def _detect_scanned_pdf(file_path: str) -> tuple[bool, str]:
    """Detect whether a PDF is a scanned (image-only) document.

    Samples up to 10 pages spread evenly across the document and measures
    average extractable text per page. A scanned PDF has negligible selectable
    text; the threshold is 100 characters/page on average.

    Args:
        file_path: Local path to the PDF file.

    Returns:
        (is_scanned, reason) tuple. is_scanned=True when the PDF is image-only.
    """
    if not file_path.lower().endswith(".pdf"):
        return False, ""
    if not Path(file_path).exists():
        return True, f"Rejected: File not found — {file_path}"
    try:
        import fitz  # PyMuPDF
        doc = fitz.open(file_path)
        total_pages = len(doc)
        if total_pages == 0:
            doc.close()
            return True, "Rejected: PDF file has 0 pages and cannot be processed."
        # Sample up to 10 pages spread across the document
        sample_size = min(10, total_pages)
        step = max(1, total_pages // sample_size)
        indices = [i * step for i in range(sample_size) if i * step < total_pages]
        total_chars = sum(len(doc[i].get_text().strip()) for i in indices)
        doc.close()
        avg_chars = total_chars / len(indices)
        if avg_chars < 100:
            return True, (
                f"Rejected: PDF appears to be a scanned document (image-only). "
                f"Average extractable text is {avg_chars:.0f} characters/page "
                f"across {len(indices)} sampled pages (threshold: 100). "
                f"Text extraction cannot process scanned PDFs — "
                f"please provide a text-based PDF instead."
            )
        return False, ""
    except Exception as e:
        return True, f"Rejected: Failed to parse PDF file — {e}. The file may be corrupted or in an unsupported format."


# NOTE: This endpoint is intentionally offline — no @router.post decorator.
# Local file upload via multipart form has been disabled.
# Use POST /from-s3 to register documents via S3 URL instead.
async def upload_file(
    dataset_id: str,
    file: UploadFile = File(...),
    metadata: str = Form(...)
):
    """Upload file to dataset."""
    # Parse metadata string (form data is always str)
    try:
        metadata_dict = json.loads(metadata)
        if not isinstance(metadata_dict, dict):
            raise ValueError("Metadata must be a JSON object")
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Metadata must be valid JSON")
    try:
        _validate_metadata(metadata_dict)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    # Save file using storage abstraction
    storage = get_storage()
    file_path = storage.save(file.file, file.filename, dataset_id)

    # Reject scanned (image-only) PDFs before creating any record
    is_scanned, reason = _detect_scanned_pdf(file_path)
    if is_scanned:
        raise HTTPException(status_code=422, detail=reason)

    # Create document record (handles dataset validation, hash, duplicate check, insert)
    try:
        result = _create_document_record(dataset_id, file_path, file.filename, metadata_dict)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    # Cache PDF to PDF_FILES_DIR for locate_pages fast lookup
    _cache_pdf_to_files_dir(file_path, result["doc_id"])

    message = "File already exists, reusing existing document" if result["is_duplicate"] else "File uploaded successfully"
    return ApiResponse(success=True, code=200, message=message, data=result)


@router.post("/from-s3", response_model=ApiResponse[dict])
async def upload_from_s3(
    dataset_id: str,
    s3_url: str = Body(..., embed=True),
    metadata: Dict[str, Any] = Body(..., embed=True),
):
    """
    Download a file from S3 and register it in the dataset.

    Identical flow to local upload: download -> compute content hash -> insert record.
    doc_id = content hash, same dedup logic applies.
    S3 existence is verified upfront via a HEAD request (400 if not found).
    """
    if not s3_url.startswith("s3://"):
        raise HTTPException(status_code=400, detail="s3_url must start with 's3://'")

    # Validate metadata structure
    try:
        _validate_metadata(metadata)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    filename = Path(s3_url).name
    if not filename:
        raise HTTPException(status_code=400, detail="Cannot determine filename from S3 URL")

    # Verify the file exists on S3 before downloading (cheap HEAD request)
    try:
        get_s3_object_info(
            s3_url,
            aws_access_key_id=config.AWS_ACCESS_KEY_ID,
            aws_secret_access_key=config.AWS_SECRET_KEY,
        )
    except FileNotFoundError:
        raise HTTPException(status_code=400, detail=f"File not found in S3: {s3_url}")
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to verify S3 file: {str(e)}")

    # Download to a temp location for hash calculation only
    storage = get_storage()
    temp_dir = Path(tempfile.gettempdir()) / f"s3_{dataset_id}"
    temp_dir.mkdir(parents=True, exist_ok=True)
    temp_file = temp_dir / filename

    try:
        await download_file_from_s3_async(
            s3_url,
            temp_file,
            aws_access_key_id=config.AWS_ACCESS_KEY_ID,
            aws_secret_access_key=config.AWS_SECRET_KEY,
        )
    except Exception as e:
        temp_file.unlink(missing_ok=True)
        raise HTTPException(status_code=500, detail=f"Failed to download file from S3: {str(e)}")

    # Reject scanned (image-only) PDFs before creating any record
    is_scanned, reason = _detect_scanned_pdf(str(temp_file))
    if is_scanned:
        temp_file.unlink(missing_ok=True)
        raise HTTPException(status_code=422, detail=reason)

    # Create document record using temp file for hash/size, but store s3_url as file_path
    result = None
    try:
        result = _create_document_record(
            dataset_id, str(temp_file), filename, metadata, s3_url=s3_url
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        print(f"[ERROR] Document record creation failed for {filename}: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to create document record: {str(e)}")
    finally:
        # Cache PDF to PDF_FILES_DIR before removing temp file
        if result:
            _cache_pdf_to_files_dir(str(temp_file), result["doc_id"])
        # Temp file is no longer needed; worker fetches directly from S3
        temp_file.unlink(missing_ok=True)

    message = (
        "File already exists, reusing existing document"
        if result["is_duplicate"]
        else "File downloaded from S3 and registered successfully"
    )
    return ApiResponse(success=True, code=200, message=message, data=result)


@router.post("/{doc_id}/tasks", response_model=ApiResponse[TaskResponse])
async def create_task(
    dataset_id: str,
    doc_id: str,
    mode: ProcessingMode = ProcessingMode.CLASSIC,
    reader: ReaderType = ReaderType.MINERU
):
    """Create a processing task for an existing document."""
    conn = get_connection()
    doc_repo = DocumentRepository(conn)
    task_repo = TaskRepository(conn)

    doc = doc_repo.get(doc_id, dataset_id)
    if not doc:
        conn.close()
        raise HTTPException(status_code=404, detail="Document not found")

    # Parse document metadata
    doc_metadata = json.loads(doc["metadata"]) if doc.get("metadata") else None

    timestamp = now()
    task_id = generate_id()

    # Create Task record
    task_repo.create(task_id, dataset_id, doc_id, mode.value, reader.value,
                     TaskStatus.PENDING, 0, timestamp)

    # Update document status and task_id
    doc_repo.update_task_link(doc_id, DocumentStatus.PROCESSING, task_id, timestamp)
    conn.close()

    # Enqueue task to Dramatiq worker
    later.process_document(task_id)

    return ApiResponse(
        success=True,
        code=200,
        message="Task created successfully",
        data=TaskResponse(
            task_id=str(task_id),
            dataset_id=dataset_id,
            doc_id=doc_id,
            mode=mode.value,
            reader=reader.value,
            status=TaskStatus.PENDING,
            progress=0,
            metadata=doc_metadata,
            created_at=timestamp,
            updated_at=timestamp
        )
    )


@router.get("/{doc_id}/units/export")
async def export_units_excel(dataset_id: str, doc_id: str):
    """Export all units for a document as an Excel file."""
    import io
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl is required: pip install openpyxl")

    # Verify document exists
    conn = get_connection()
    row = DocumentRepository(conn).get(doc_id, dataset_id)
    conn.close()
    if not row:
        raise HTTPException(status_code=404, detail="Document not found")

    # Fetch units via internal units router logic
    from .units import _list_units_for_doc
    units = await _list_units_for_doc(dataset_id, doc_id)

    # Build Excel in memory
    COLUMNS = [
        ("unit_id",           "Unit ID",           38),
        ("unit_type",         "Type",              10),
        ("has_views",         "LOD?",               8),
        ("content",           "Content",           80),
        ("embedding_content", "Embedding Content", 80),
    ]
    HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    CELL_FILL   = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True)
    WRAP        = Alignment(wrap_text=True, vertical="top")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Units"

    for col_idx, (key, label, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row_idx, unit in enumerate(units, 2):
        row_data = {
            "unit_id":           unit.get("unit_id", ""),
            "unit_type":         unit.get("unit_type", ""),
            "has_views":         "YES" if unit.get("has_views") else "",
            "content":           unit.get("content", ""),
            "embedding_content": unit.get("embedding_content", "") or "",
        }
        for col_idx, (key, label, width) in enumerate(COLUMNS, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(key, ""))
            cell.alignment = WRAP
            cell.fill = CELL_FILL

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"units_{doc_id[:8]}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{doc_id}/download")
async def download_document(dataset_id: str, doc_id: str):
    """Download the original file for a document.

    Works for both local and S3-backed files.
    """
    conn = get_connection()
    row = DocumentRepository(conn).get(doc_id, dataset_id)
    conn.close()

    if not row:
        raise HTTPException(status_code=404, detail="Document not found")

    file_name = row["file_name"]
    file_path = row["file_path"]

    if file_path.startswith("s3://"):
        # Stream from S3
        import io
        from ..utils.s3 import download_file_from_s3
        buf = io.BytesIO()
        try:
            import boto3
            import re
            m = re.match(r"s3://([^/]+)/(.+)", file_path)
            if not m:
                raise HTTPException(status_code=500, detail="Invalid S3 URL")
            bucket, key = m.group(1), m.group(2)
            s3 = boto3.client(
                "s3",
                aws_access_key_id=config.AWS_ACCESS_KEY_ID,
                aws_secret_access_key=config.AWS_SECRET_KEY,
            )
            s3.download_fileobj(bucket, key, buf)
            buf.seek(0)
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Failed to download from S3: {e}")
        return StreamingResponse(
            buf,
            media_type="application/octet-stream",
            headers={"Content-Disposition": f'attachment; filename="{file_name}"'},
        )
    else:
        # Local file
        p = Path(file_path)
        if not p.is_absolute():
            p = (config.RAG_SERVICE_DIR / p).resolve()
        if not p.exists():
            raise HTTPException(status_code=404, detail="File not found on disk")
        return FileResponse(
            path=str(p),
            filename=file_name,
            media_type="application/octet-stream",
        )


@router.get("/{doc_id}/views", response_model=ApiResponse[list])
def get_document_views(
    dataset_id: str,
    doc_id: str,
    level: Optional[str] = None,
):
    """Get LOD views for a document.

    Args:
        level: Filter by level - low | medium | high | full. Returns all if omitted.

    Response items:
        - LOW/MEDIUM/FULL: {level, content, token_count}
        - HIGH (DocTree):  {level, tree, token_count}
    """
    from zag.storages.vector import QdrantVectorStore
    from zag.embedders import Embedder
    from zag.schemas import LODLevel

    # Validate level param
    valid_levels = {lv.value for lv in LODLevel}
    if level and level not in valid_levels:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid level '{level}'. Must be one of: {sorted(valid_levels)}"
        )

    # Resolve collection_name
    conn = get_connection()
    ds = DatasetRepository(conn).get(dataset_id)
    conn.close()

    if not ds:
        raise HTTPException(status_code=404, detail="Dataset not found")
    else:
        collection_name = ds["name"]

    # Fetch LOD unit via QdrantVectorStore.fetch
    try:
        embedder = Embedder(config.EMBEDDING_URI, api_key=config.OPENAI_API_KEY)
        vector_store = QdrantVectorStore.server(
            host=config.VECTOR_STORE_HOST,
            port=config.VECTOR_STORE_PORT,
            grpc_port=config.VECTOR_STORE_GRPC_PORT,
            prefer_grpc=True,
            collection_name=collection_name,
            embedder=embedder,
        )
        units = vector_store.fetch({
            "doc_id": doc_id,
            "metadata.custom.mode": "lod",
        })
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to fetch views: {str(e)}")

    if not units:
        raise HTTPException(status_code=404, detail="No LOD views found for this document")

    # Use the first (and only) LOD unit
    lod_unit = units[0]
    if not lod_unit.views:
        raise HTTPException(status_code=404, detail="LOD unit has no views")

    # Build response items, optionally filtered by level
    result = []
    tags = _extract_tags(lod_unit)
    for view in lod_unit.views:
        if level and view.level != level:
            continue

        if view.level == LODLevel.HIGH:
            result.append({
                "level": view.level if isinstance(view.level, str) else view.level.value,
                "tree": view.content,
                "token_count": view.token_count,
                "tags": tags,
            })
        else:
            result.append({
                "level": view.level if isinstance(view.level, str) else view.level.value,
                "content": view.content,
                "token_count": view.token_count,
                "tags": tags,
            })

    return ApiResponse(success=True, code=200, data=result)


@router.get("/{doc_id}/tasks", response_model=ApiResponse[List[TaskResponse]])
def list_document_tasks(dataset_id: str, doc_id: str):
    """Get all tasks for a specific document."""
    conn = get_connection()
    doc_repo = DocumentRepository(conn)
    task_repo = TaskRepository(conn)

    if not doc_repo.get(doc_id, dataset_id):
        conn.close()
        raise HTTPException(status_code=404, detail="Document not found")

    rows = task_repo.list_by_doc(dataset_id, doc_id)
    conn.close()

    results = []
    for row in rows:
        error_message = json.loads(row["error_message"]) if row["error_message"] else None
        doc_metadata = json.loads(row["doc_metadata"]) if row.get("doc_metadata") else None
        results.append(TaskResponse(
            task_id=str(row["id"]),
            dataset_id=str(row["dataset_id"]),
            doc_id=str(row["doc_id"]),
            mode=row.get("mode", "classic"),
            reader=row.get("reader", "mineru"),
            status=row["status"],
            progress=row["progress"],
            metadata=doc_metadata,
            error_message=error_message,
            created_at=row["created_at"],
            updated_at=row["updated_at"]
        ))

    return ApiResponse(success=True, code=200, data=results)


@router.get("", response_model=ApiResponse[List[DocumentResponse]])
def list_documents(
    dataset_id: str,
    status: Optional[str] = None,
    limit: Optional[int] = None,
):
    """List documents in a dataset."""
    conn = get_connection()
    dataset_repo = DatasetRepository(conn)
    doc_repo = DocumentRepository(conn)

    if not dataset_repo.exists(dataset_id):
        conn.close()
        raise HTTPException(status_code=404, detail="Dataset not found")

    rows = doc_repo.list_by_dataset(dataset_id, status)
    conn.close()

    if limit is not None:
        rows = rows[:limit]

    # Pre-compute base_dir_str once for all file_url builds (pure string, no I/O per row)
    base_dir_str = Path(get_storage().base_dir).resolve().as_posix()

    results = []
    for row in rows:
        doc_metadata = json.loads(row["metadata"]) if row.get("metadata") else None
        results.append(DocumentResponse(
            doc_id=str(row["id"]),
            dataset_id=str(row["dataset_id"]),
            file_name=row["file_name"],
            file_path=row["file_path"],
            file_url=_build_file_url(row["file_path"], base_dir_str),
            workspace_dir=row["workspace_dir"],
            file_size=row["file_size"],
            file_type=row["file_type"],
            file_hash=row.get("file_hash"),
            metadata=doc_metadata,
            status=row["status"],
            task_id=str(row["task_id"]) if row["task_id"] else None,
            unit_count=row["unit_count"],
            created_at=row["created_at"],
            updated_at=row["updated_at"]
        ))

    return ApiResponse(success=True, code=200, data=results)


@router.get("/{doc_id}", response_model=ApiResponse[DocumentResponse])
async def get_document(
    dataset_id: str,
    doc_id: str,
    realtime_count: bool = Query(False, description="Fetch live unit count from vector DB instead of cached DB value"),
):
    """Get document by ID."""
    conn = get_connection()
    row = DocumentRepository(conn).get(doc_id, dataset_id)
    conn.close()

    if not row:
        raise HTTPException(status_code=404, detail="Document not found")

    # Parse document metadata
    doc_metadata = json.loads(row["metadata"]) if row.get("metadata") else None

    # Build file_url for the worker to fetch the file
    file_url = _build_file_url(row["file_path"])

    # Unit count: use cached DB value by default; only hit Qdrant when explicitly requested.
    # Calling acount on every GET is risky: a broken gRPC/HTTP connection will hang
    # the request until timeout and pollute all callers sharing the singleton cache.
    unit_count = row.get("unit_count")
    if realtime_count:
        collection_name = None
        try:
            from .query import get_dataset_info, _get_vector_store, _invalidate_vector_store
            collection_name, engine = get_dataset_info(dataset_id)
            if engine == "qdrant":
                vector_store = _get_vector_store(collection_name)
                unit_count = await asyncio.wait_for(
                    vector_store.acount({"doc_id": doc_id}),
                    timeout=10.0,
                )
        except asyncio.TimeoutError:
            if collection_name:
                _invalidate_vector_store(collection_name)
            import logging
            logging.getLogger(__name__).warning(
                "acount timed out for doc %s, vector store cache evicted", doc_id
            )
        except Exception as e:
            if collection_name:
                _invalidate_vector_store(collection_name)
            import logging
            logging.getLogger(__name__).warning("acount failed: %s", e, exc_info=True)

    data = DocumentResponse(
        doc_id=str(row["id"]),
        dataset_id=str(row["dataset_id"]),
        file_name=row["file_name"],
        file_path=row["file_path"],
        file_url=file_url,
        workspace_dir=row["workspace_dir"],
        file_size=row["file_size"],
        file_type=row["file_type"],
        file_hash=row.get("file_hash"),
        metadata=doc_metadata,
        status=row["status"],
        task_id=str(row["task_id"]) if row["task_id"] else None,
        unit_count=unit_count,
        created_at=row["created_at"],
        updated_at=row["updated_at"]
    )

    return ApiResponse(success=True, code=200, data=data)


@router.patch("/{doc_id}", response_model=ApiResponse[DocumentResponse])
async def update_document(
    dataset_id: str,
    doc_id: str,
    metadata: Dict[str, Any] = Body(..., embed=True)
):
    """Update document metadata."""
    # Validate metadata structure
    try:
        _validate_metadata(metadata)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    conn = get_connection()
    doc_repo = DocumentRepository(conn)

    row = doc_repo.get(doc_id, dataset_id)
    if not row:
        conn.close()
        raise HTTPException(status_code=404, detail="Document not found")

    timestamp = now()
    metadata_json = json.dumps(metadata)
    doc_repo.update_metadata(doc_id, metadata_json, timestamp)

    updated = doc_repo.get(doc_id)
    conn.close()

    doc_metadata = json.loads(updated["metadata"]) if updated["metadata"] else None

    return ApiResponse(
        success=True,
        code=200,
        message="Document updated successfully",
        data=DocumentResponse(
            doc_id=str(updated["id"]),
            dataset_id=str(updated["dataset_id"]),
            file_name=updated["file_name"],
            file_path=updated["file_path"],
            workspace_dir=updated["workspace_dir"],
            file_size=updated["file_size"],
            file_type=updated["file_type"],
            file_hash=updated.get("file_hash"),
            metadata=doc_metadata,
            status=updated["status"],
            task_id=str(updated["task_id"]) if updated["task_id"] else None,
            unit_count=updated["unit_count"],
            created_at=updated["created_at"],
            updated_at=updated["updated_at"]
        )
    )


@router.delete("/{doc_id}", response_model=ApiResponse[MessageResponse])
def delete_document(dataset_id: str, doc_id: str):
    """Delete document and cleanup vector store.
    
    Returns 409 Conflict if the document has active dependencies.
    Callers should resolve dependencies before retrying.
    """
    conn = get_connection()
    doc_repo = DocumentRepository(conn)
    dep_repo = DependencyRepository(conn)
    task_repo = TaskRepository(conn)
    dataset_repo = DatasetRepository(conn)

    doc = doc_repo.get(doc_id, dataset_id)
    if not doc:
        conn.close()
        raise HTTPException(status_code=404, detail="Document not found")

    # Check for active dependencies (documents that depend on this one)
    dependencies = dep_repo.list_by_target(doc_id, dataset_id)
    if dependencies:
        conn.close()
        dep_list = [{"rule": d["rule"], "target_doc_id": d["target_doc_id"]} for d in dependencies]
        raise HTTPException(
            status_code=409,
            detail={
                "message": "Document has active dependencies",
                "dependencies": dep_list
            }
        )

    ds = dataset_repo.get(dataset_id)
    if not ds:
        conn.close()
        raise HTTPException(status_code=404, detail="Dataset not found")

    collection_name = ds["name"]

    # Delete from database: source-level dependencies, tasks, document
    dep_repo.delete_by_rule(str(Rule.build(DependencySource.DOC, doc_id)), dataset_id)
    task_repo.delete_by_doc(doc_id)
    doc_repo.delete(doc_id)
    conn.close()
    
    # Cleanup vector store
    try:
        from zag.storages.vector import QdrantVectorStore
        from zag.embedders import Embedder
        
        embedder = Embedder(
            config.EMBEDDING_URI,
            api_key=config.OPENAI_API_KEY
        )
        vector_store = QdrantVectorStore.server(
            host=config.VECTOR_STORE_HOST,
            port=config.VECTOR_STORE_PORT,
            grpc_port=config.VECTOR_STORE_GRPC_PORT,
            prefer_grpc=True,
            collection_name=collection_name,
            embedder=embedder
        )
        vector_store.remove({"doc_id": doc_id})
    except Exception as e:
        # Log but don't fail the API call
        print(f"Warning: Failed to cleanup vector store for doc_id {doc_id}: {e}")

    # Cleanup fulltext store (Meilisearch)
    try:
        from zag.indexers import FullTextIndexer

        fulltext_indexer = FullTextIndexer(
            url=config.MEILISEARCH_HOST,
            index_name=collection_name,
            api_key=config.MEILISEARCH_API_KEY,
            auto_create_index=False,
        )
        deleted = fulltext_indexer.delete_by_doc_id(doc_id)
        if deleted:
            print(f"Cleaned up {deleted} units from Meilisearch for doc_id {doc_id}")
    except Exception as e:
        print(f"Warning: Failed to cleanup Meilisearch for doc_id {doc_id}: {e}")
    
    return ApiResponse(
        success=True,
        code=200,
        message="Document deleted successfully",
        data=MessageResponse(message="Document deleted successfully")
    )


# TEMPORARY HACK — TODO: delete this function once the two broken docs are re-indexed.
# Context: doc IDs below have a malformed page-1 that catches too many fuzzy matches,
# producing false-positive page-1 results. Forcing those results to "not found" is the
# quickest safe workaround while waiting for a proper re-parse.
_BAD_PAGE1_DOC_IDS = {"883489afd016d9e4", "86ffa34b6ce3cecc"}


def _patch_bad_page1_docs(results: list["LocatePageResult"]) -> list["LocatePageResult"]:
    """TEMPORARY: Suppress false-positive page-1 hits for known broken documents.

    For doc IDs in ``_BAD_PAGE1_DOC_IDS``, any result whose only matched page is
    page 1 is replaced with a not-found result. Multi-page results that include
    page 1 alongside other pages are left untouched.

    TODO: Remove this function (and its call site) once the affected documents
    have been re-indexed with a corrected PDF parser.
    """
    patched = []
    for r in results:
        if r.doc_id in _BAD_PAGE1_DOC_IDS and 1 in r.page_numbers:
            patched.append(r.model_copy(update={
                "page_numbers": [],
                "found": False,
                "error": "Text not found in source document",
            }))
        else:
            patched.append(r)
    return patched


@router.post("/locate", response_model=ApiResponse[LocatePageResponse])
async def locate_pages(
    dataset_id: str,
    request: LocatePageRequest = Body(...),
):
    """
    Locate page numbers for text snippets.
    
    Batch API that accepts multiple items, each with doc_id + text_start (+ optional text_end).
    Returns page numbers for each item.
    
    Performance:
    - Concurrent PDF loading for different docs
    - Concurrent text search within each PDF
    """
    import asyncio
    from collections import defaultdict
    from concurrent.futures import ThreadPoolExecutor
    
    def get_file_path_for_doc(doc_id: str) -> str | None:
        """
        Return a local path to the original document for the given doc_id.

        Lookup order:
        1. PDF_FILES_DIR/{doc_id}*.<known-ext>  – pick the newest by mtime
        2. DB file_path: local path (absolute or relative to rag-service dir)
        3. S3 download → saved to PDF_FILES_DIR/{doc_id}{original_ext} for next time

        Never raises – returns None on any failure so callers can fallback safely.
        """
        _DOC_EXTENSIONS = {
            '.pdf', '.docx', '.doc', '.md', '.txt', '.markdown',
            '.pptx', '.ppt', '.xlsx', '.xls',
        }
        try:
            from ..utils.s3 import download_file_from_s3

            pdf_files_dir = Path(config.PDF_FILES_DIR)

            # 1. Find all local variants matching {doc_id}*, pick newest by mtime
            if pdf_files_dir.exists():
                candidates = sorted(
                    [
                        p for p in pdf_files_dir.glob(f"{doc_id}*")
                        if p.suffix.lower() in _DOC_EXTENSIONS
                    ],
                    key=lambda p: p.stat().st_mtime,
                    reverse=True,
                )
                if candidates:
                    return str(candidates[0])

            # 2. Get file_path from DB
            conn = get_connection()
            try:
                row = DocumentRepository(conn).get(doc_id)
                if not row or not row["file_path"]:
                    return None
                file_path = row["file_path"]
            finally:
                conn.close()

            if file_path.startswith("s3://"):
                # 3. Download from S3, preserve original extension from S3 path
                pdf_files_dir.mkdir(parents=True, exist_ok=True)
                s3_ext = Path(file_path).suffix or '.pdf'
                local_file = pdf_files_dir / f"{doc_id}{s3_ext}"
                ok = download_file_from_s3(file_path, local_file)
                return str(local_file) if ok else None
            else:
                # Local path: try as-is first, then relative to rag-service dir
                p = Path(file_path)
                if p.is_absolute():
                    return str(p) if p.exists() else None
                # Relative path – resolve against the rag-service directory
                resolved = (config.RAG_SERVICE_DIR / p).resolve()
                return str(resolved) if resolved.exists() else None
        except Exception:
            return None

    def process_single_doc(doc_id: str, items: list) -> list[LocatePageResult]:
        """Process all items for a single document. Uses diskcache when available. Never raises."""
        try:
            cache = _get_locate_cache()

            # Step 1: get (full_text, page_positions) — from cache or by parsing the PDF.
            # On a cache miss, parse via fitz and write back to cache before searching,
            # so all searches always run on the same cached-text path.
            entry = cache.get(doc_id) if cache is not None else None

            if entry is None:
                file_path = get_file_path_for_doc(doc_id)
                if not (file_path and Path(file_path).suffix.lower() == '.pdf'):
                    return [LocatePageResult(request_id=item.request_id, doc_id=doc_id,
                                             page_numbers=[], found=False,
                                             error="Source file not available or not a PDF")
                            for item in items]
                entry = _extract_pdf_text_and_positions(file_path)
                if entry is None:
                    return [LocatePageResult(request_id=item.request_id, doc_id=doc_id,
                                             page_numbers=[], found=False,
                                             error="PDF appears to be scanned or contains no extractable text")
                            for item in items]
                if cache is not None:
                    try:
                        cache.set(doc_id, entry)
                    except Exception:
                        pass

            full_text, page_positions = entry
            results = []
            for item in items:
                try:
                    page_numbers, found = _search_in_text(
                        full_text, page_positions, item.text_start, item.text_end
                    )
                    results.append(LocatePageResult(
                        request_id=item.request_id,
                        doc_id=doc_id,
                        page_numbers=page_numbers or [],
                        found=found,
                        error=None if found else "Text not found in source document",
                    ))
                except Exception as e:
                    results.append(LocatePageResult(
                        request_id=item.request_id,
                        doc_id=doc_id,
                        page_numbers=[],
                        found=False,
                        error=f"Search error: {str(e)}",
                    ))
            return results

        except Exception as e:
            return [
                LocatePageResult(
                    request_id=item.request_id,
                    doc_id=doc_id,
                    page_numbers=[],
                    found=False,
                    error=f"Doc processing error: {str(e)}",
                )
                for item in items
            ]

    # Group items by doc_id, then process all docs concurrently
    doc_groups: dict[str, list] = defaultdict(list)
    for item in request.items:
        doc_groups[item.doc_id].append(item)

    loop = asyncio.get_running_loop()

    # Cache path: no disk I/O per-doc, CPU-only fuzzy search → can safely use more workers.
    # Fitz fallback still benefits from parallelism without heavy I/O contention.
    max_doc_workers = min(32, len(doc_groups))

    with ThreadPoolExecutor(max_workers=max_doc_workers) as executor:
        futures = [
            loop.run_in_executor(executor, process_single_doc, doc_id, items)
            for doc_id, items in doc_groups.items()
        ]
        all_results = await asyncio.gather(*futures)
    
    # Flatten results
    results = []
    for doc_results in all_results:
        results.extend(doc_results)

    # TEMPORARY HACK — TODO: remove once upstream PDF indexing is fixed for these two docs.
    # These documents have a bad page-1 extraction (likely a cover/TOC page that absorbs
    # fuzzy matches). Any result landing on page 1 is almost certainly a false positive,
    # so we force them to "not found" to avoid surfacing garbage to the caller.
    results = _patch_bad_page1_docs(results)

    return ApiResponse(
        success=True,
        code=200,
        data=LocatePageResponse(results=results)
    )
